import requests
from bs4 import BeautifulSoup
from pymongo import MongoClient
from openpyxl import load_workbook  # 엑셀을 불러오기위해 openpyxl 패키지 사용

client = MongoClient('localhost', 27017)
db = client.vegan
collection = db['vegan']
db_selections = collection.find({})

wb = load_workbook(filename='VeganMap_csv.xlsx')
# 엑셀 불러오기
ws = wb.active
# 현재 활성화 되어있는 시트 선택! [ 저희는 시트가 하나라서 그게 선택됩니다!]

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari / 537.36'}
count = 2  # 첫번째 행은 header로 값을 넣는 공간이 아니라서 2부터 시작하게 count로 설정!

for db in db_selections:
    shop_region = db['자치구']
    shop_name = db['상호명']
    search = shop_region + ' ' + shop_name

    data = requests.get(
        f'https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query={search}',
        headers=headers)

    soup = BeautifulSoup(data.text, 'html.parser')
    if data.status_code == 200:
        try:
            images = soup.select('#place_main_ct')
            if len(images) != 0:
                for image in images:
                    img_url = \
                        image.select_one(
                            'div > section > div > div.top_photo_area.type_v4 > div:nth-child(1) > a > img')[
                            'src']
                    ws.cell(row=count, column=10).value = img_url
                    print(f'{search} 성공')
                    count += 1
            else:
                ws.cell(row=count, column=10).value = f"이미지가 없습니다."
                print(f'{search} 실패')
                count += 1
        except Exception as e:
            ws.cell(row=count, column=10).value = f"이미지가 없습니다."
            print(f'{search} 실패')
            count += 1
wb.save(filename='VeganMap.xlsx')
