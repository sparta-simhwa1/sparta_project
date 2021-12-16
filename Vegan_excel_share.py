from openpyxl import load_workbook  # 엑셀을 불러오기위해 openpyxl 패키지 사용
import urllib.request
import json

wb = load_workbook(filename='VeganMap.xlsx')
# 엑셀 불러오기
ws = wb.active
# 현재 활성화 되어있는 시트 선택! [ 저희는 시트가 하나라서 그게 선택됩니다!]
temp = []  # 엑셀에 저장된 정보를 다 저장! (나중에 어떤 기능을 추가할지 몰라서 일단 다 저장하는걸 넣어놨어요!)
addresses = []  # 엑셀 내용중 상세주소만 저장! ( 위도 경도 변환을 위해! )

for i in range(2, ws.max_row + 1):  # ws.max_row는 모든 행의 길이를 반환! for문으로 모두 돌아야하기 때문에 +1
    temp.append([]) # 이중 리스트 구현을 위해 리스트를 넣어서 공간 만들기!
    for cell in ws[i]:
        temp[-1].append(cell.value)

for i in range(0, ws.max_row - 1):
    addresses.append(temp[i][5])

count = 2 # 첫번째 행은 header로 값을 넣는 공간이 아니라서 2부터 시작하게 count로 설정!

for address in addresses:
    encoding_address = urllib.parse.quote_plus(address)
    # url을 읽을 수 있게 ascii 값으로 변환을 위해 urllib.parse.quote_plus 사용하기
    url = f'https://naveropenapi.apigw.ntruss.com/map-geocode/v2/geocode?query={encoding_address}'
    request = urllib.request.Request(url)
    request.add_header("X-NCP-APIGW-API-KEY-ID", Client ID) # Client ID 키를 문자열 형태로 넣기
    request.add_header("X-NCP-APIGW-API-KEY", Client Secret) # Client Secret 키를 문자열 형태로 넣기

    response = urllib.request.urlopen(request)
    response_code = response.getcode() # 불러온 사이트의 상태가 정상인지 확인

    if response_code == 200: #정상 일시 시행
        try:
            response_body = response.read()
            # 네이버 Open API를 통해서 수신된 데이터가 JSON 포멧이기 때문에,
            # JSON 포멧 데이터를 파싱해서 사전데이터로 만들어주는 json 라이브러라를 사용
            data = json.loads(response_body)
            ws.cell(row=count, column=8).value = data['addresses'][0]['y']
            ws.cell(row=count, column=9).value = data['addresses'][0]['x']
            print(ws.cell(row=count, column=6).value)  # 경도와 위도에 해당하는 x, y값을 출력
            count += 1 # 행 아래로 내리기 위해 count 값 추가
        except:
            ws.cell(row=count, column=8).value = 'error'
            ws.cell(row=count, column=9).value = 'error'
            count += 1
            print(f'error : {count}')

wb.save(filename='VeganMap.xlsx')
